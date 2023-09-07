import openpyxl


def convert_name(name):
    parts = name.split(", ")

    if len(parts) == 2:
        last_name, first_name = parts
        converted_name = f"{first_name} {last_name}"
        return converted_name
    else:
        return name


def read_booklist():
    wb = openpyxl.load_workbook("master.xlsx")
    sheet = wb["Accession Details"]

    result_dict = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        details = dict()
        details["accession_number"] = row[0]
        details["author"] = convert_name(row[3])
        details["title"] = row[4]
        details["call_number"] = row[5]
        details["publisher"] = row[6]
        details["place_of_publication"] = row[7]
        details["isbn"] = row[8]
        details["vendor"] = row[9]
        details["bill_number"] = row[10]
        details["price"] = row[12]
        result_dict.append(details)

    return result_dict


def read_namelist():
    wb = openpyxl.load_workbook("master.xlsx")
    sheet = wb["Issue Details"]

    result_list = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        result_list.append((row[1], row[2]))

    return result_list

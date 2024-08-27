import openpyxl
from shelfmaster.models import User, Entity
from shelfmaster import app, db


def convert_name(name):
    try:
        parts = name.split(", ")
    except AttributeError:
        return name

    if len(parts) == 2:
        last_name, first_name = parts
        converted_name = f"{first_name} {last_name}"
        return converted_name
    else:
        return name


def read_namelist_from_upload(filename):
    workbook = openpyxl.load_workbook(filename)

    sheet = workbook["1"]
    result_dicts = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        details = dict()

        details["roll_number"] = row[0]
        details["usn"] = row[1]
        details["name"] = row[2]
        details["class_section"] = row[3]

        if all(details.values()):
            result_dicts.append(details)

    return result_dicts


def read_booklist_from_upload(filename):
    workbook = openpyxl.load_workbook(filename)

    sheet = workbook["1"]
    result_dicts = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        details = dict()
        details["accession_number"] = row[0]
        details["author"] = row[1]
        details["title"] = row[2]
        details["call_number"] = row[3]
        details["publisher"] = row[4]
        details["isbn"] = row[5]
        details["vendor"] = row[6]
        details["bill_number"] = row[7]
        details["bill_date"] = row[8]
        details["price"] = row[9]
        details["remarks"] = row[10]
        details["language"] = row[11]
        if any(details.values()):
            result_dicts.append(details)

    return result_dicts
